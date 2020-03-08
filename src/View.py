#!/usr/bin/env python
# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import re
from Select import *

#currentdb = 'test'

def ViewCreate(currentdb, viewname, sql):
    wb = load_workbook('data/view.xlsx')
    try:
        sheet = wb[viewname]
    except:
        sheet = wb.create_sheet()
        sheet.title = viewname
    sql = sql.lower()
    sqlItem = re.split(r'[, ]', sql)
    RemoveKong(sqlItem)
    headers = sqlItem[sqlItem.index('select')+1 : sqlItem.index('from')]
    #print headers
    viewtableItem = Select(sql, currentdb)
    for i in xrange(1, len(headers)+1):
        sheet.cell(1, i).value = headers[i-1]
        #print sheet.cell(1, i).value

    for i in xrange(2, len(viewtableItem)+1):
        for j in xrange(1, len(headers)+1):
            sheet.cell(i,j).value = viewtableItem[i-2][j-1]
    wb.save('data/view.xlsx')
    print '%s View has been create.' % viewname

# sql = 'Select id course grade From aaa'
# ViewCreate(currentdb, 'IdGrade', sql)

