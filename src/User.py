#!/usr/bin/env python
# -*- coding: utf-8 -*-

import hashlib
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from Select import *
import re

def md5(ss):  # 密码使用md5加密，提高安全性
    m = hashlib.md5()
    m.update(ss)
    return m.hexdigest()

def Login(user, passwd):
    wb = load_workbook('data/user.xlsx')
    sheet = wb.active
    ret = []  # ret 用于保存是否是管理员、操作的数据库、所具有的权限
    for i in xrange(2, sheet.max_row+1):
        if sheet.cell(row=i, column=1).value == user:
            if sheet.cell(row=i, column=2).value == md5(passwd):
                for j in xrange(3, sheet.max_column + 1):
                    ret.append(sheet.cell(row=i, column=j).value)
                # print ret  # 测试获取的权限
                print "Login Successfully.%s,Welcome!" % user
                return ret
            else:
                print "your password is wrong."
                exit(0)
    print "user not exists"
    exit(0)

def Grant(sql):
    sql = sql.lower()
    sqlItem = re.split(r'[, ]', sql)
    RemoveKong(sqlItem)
    # print sqlItem

    privilege = sqlItem[sqlItem.index('grant')+1: sqlItem.index('on')]
    dbname = sqlItem[sqlItem.index('on')+1: sqlItem.index('to')]
    username = sqlItem[sqlItem.index('user')+1]

    # print privilege
    # print dbname
    # print username

    wb = load_workbook('data/user.xlsx')
    ws = wb.active
    for i in xrange(2, ws.max_row + 1):
        if ws.cell(row=i, column=1).value == username:
            if ws.cell(row=i, column=3).value is True:
                print 'The user:%s is an administrator and does not need to modify the permissions' % username
                return [0]
            else:
                if ws.cell(row=i, column=4).value == 'all':
                    print 'The user %s can use all database.' % username
                else:
                    havingdb = ws.cell(i,4).value.split(',')
                    for d in dbname:
                        if d in havingdb:
                            print 'already can use %s' % d
                        else:
                            havingdb.append(d)
                    havingstr = ','.join(havingdb)
                    print havingstr
                    ws.cell(row=i, column=4).value = havingstr

                if ws.cell(i,5).value != 'all':
                    having = ws.cell(row=i, column=5).value.split(',')
                    for p in privilege:
                        if p in having:
                            print 'already have privilege %s' % p
                        else:
                            having.append(p)
                    havingstr = ','.join(having)
                    print havingstr
                    ws.cell(row=i, column=5).value = havingstr
                    wb.save('data/user.xlsx')
                    print 'grant sucess'
                    pri = []  # 返回权限列表
                    for j in xrange(3, ws.max_column + 1):
                        pri.append(ws.cell(row=i, column=j).value)
                    return pri
    print 'user %s not exist.' % username
    return [0]

def Revoke(sql):
    sql = sql.lower()
    sqlItem = re.split(r'[, ]', sql)
    RemoveKong(sqlItem)
    # print sqlItem

    privilege = sqlItem[sqlItem.index('revoke') + 1: sqlItem.index('on')]
    dbname = sqlItem[sqlItem.index('on') + 1: sqlItem.index('from')]
    username = sqlItem[sqlItem.index('user') + 1]

    # print privilege
    # print dbname
    # print username
    wb = load_workbook('data/user.xlsx')
    ws = wb.active
    for i in xrange(2, ws.max_row + 1):
        if ws.cell(row=i, column=1).value == username:
            if ws.cell(row=i, column=3).value is True:
                print 'The user:%s is an administrator and does not need to modify the permissions' % username
                return [0]
            else:
                having = ws.cell(row=i, column=5).value.split(',')
                for p in privilege:
                    if p not in having:
                        print 'user %s dosn\'t have privilege %s' % username % p
                    else:
                        having.remove(p)
                havingstr = ','.join(having)
                print havingstr
                ws.cell(row=i, column=5).value = havingstr
                wb.save('data/user.xlsx')
                print 'revoke sucess'
                pri = []  # 返回权限列表
                for j in xrange(3, ws.max_column + 1):
                    pri.append(ws.cell(row=i, column=j).value)
                return pri
    print 'user %s not exist.' % username
    return [0]


# sql = 'grant select, insert, update, delete on db to user test'
# Grant(sql)
# sql = 'revoke insert, update, delete on db from user test'
# Revoke(sql)

def Unique(ws, col, key):
    for i in xrange(1, ws.max_row+1):
        if ws.cell(i,col).value == key:
            return False
    return True

def Dividelist(valist, num):
    li = []
    for i in xrange(0, len(valist)/num):
        tmp_li = []
        for j in xrange(0, num):
            tmp_li.append(valist[i*num+j])
        li.append(tmp_li)
    return li

# 对参照完整性约束等做检查和处理
# ws 为要插入更新的表格
def Checking(ws, checklist, values):
    for i in xrange(len(checklist)):
        if 'not null' in checklist[i]:
            if values[i] is None or values[i]=='null':
                print 'column %s must be not null' % key
                return False

        if 'check' in checklist[i]:
            sec = checklist[i].find(')')
            for j in checklist[i]:
                if j == '>' or j == '<' or j == '=':
                    fuhao = checklist[i].find(j)
            checkval = values[i] + checklist[i][fuhao : sec]
            print checkval
            if eval(checkval) is not True:
                print '%s must be satisfied.' % checkval
                return False

        if 'unique' in checklist[i]:
            if Unique(ws, i+1, values[i]) is not True:
                print '%s must be unique.' % (checklist[i].split())[0]
                return False

        if 'primary' in checklist[i]:
            if Unique(ws, i+1, values[i]) is not True or values[i] is None or values[i]=='null':
                print '%s must be primary key.' % (checklist[i].split())[0]
                return False
    return True

# 返回用户的权限列表
def Checkprivilege(username, currentdb):
    userwb = load_workbook('data/user.xlsx')
    ws = userwb.active
    qx = []
    for r in xrange(2, ws.max_row+1):
        if ws.cell(r, 1).value == username:
            for c in xrange(3, ws.max_column+1):
                qx.append(ws.cell(r,c).value)
    #print qx
    return qx

Checking('aaa',['id int check(id>10)'], ['1'])
wb = load_workbook('data/test.xlsx')
ws = wb['aaa']
# Checking(ws, ['id int primary'], [None])

Checkprivilege('ldl','test')


