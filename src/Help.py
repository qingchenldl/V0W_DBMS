#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter

def HELP(cmd):
    if cmd == '-h':
        Help_h()
    elif cmd == 'help':
        help_all()
#   elif cmd[1] == 'databse':

def Help_h():
    print 'List of all V0W_DBMS commands:\n'
    print '-h             Display this help.'
    print 'use            Use another database. Takes database name as argument.'
    print 'help           show all databses.'
    print 'help database  Takes database name as argument.Display all tables, views, indexes.'
    print 'help table     Takes table name as argument.Display detailed information about all attributes.'
    print 'help view      Takes view name as argument.Display definition statement of view.'
    print 'help index     Takes view name as argument.Display index details.'

# 输出所有数据库
def help_all():
    try:
        print 'Database:'
        file = open('../data/DB.txt', 'r')
        alldb = file.readlines()
        for i in alldb:
            i = i.strip('\n')
            print i
    except:
        print 'Error: can\'t open DB.txt'

# 输出数据库的表和视图
def help_database(currentdb):
    if currentdb == '':
        print 'Error: No Database.'
        return 1
    wb = load_workbook('data/'+currentdb+'.xlsx')
    ws = wb['Sheet']
    print 'Table:'
    table_name = []
    for i in xrange(2, ws.max_row+1):
        if ws.cell(row=i, column=1).value is not None:
            table_name.append(ws.cell(row=i, column=1).value)
    for i in xrange(len(table_name)):
        print table_name[i],

    print ''
    print 'View:'
    try:
        file = open('data/'+currentdb + 'View.txt')
        for s in file:
            print s.split(':')[0],
        print ''
    except:
        print 'View not exist.'
    print 'Index:'
    try:
        indextab = wb['index']
        for i in xrange(1, indextab.max_row+1):
            print indextab.cell(i,1).value
    except:
        print 'index has not create.'

def help_table(currentdb, name):
    if currentdb == '':
        print 'Error:No Database'
        return 1
    wb = load_workbook('data/' + currentdb + '.xlsx')
    ws = wb['Sheet']
    for i in xrange(1, ws.max_row+1):
        if ws.cell(row=i, column=1).value == name:
            for j in xrange(1, ws.max_column+1):
                if ws.cell(row=i, column=j).value is not None:
                    print ws.cell(row=i, column=j).value
            break

def help_view(name, currentdb):
    if currentdb == '':
        print 'Error:No Database'
        return 1
    try:
        file = open('data/'+currentdb + 'View.txt')
        for s in file:
            if name in s:
                print s
                break
        file.close()
    except:
        print 'Error:No view.'

def help_index(index, currentdb):
    if currentdb == '':
        print 'Error:No Database'
        return 1
    w = load_workbook('INDEX.xlsx')
    ishave = False
    for name in w.sheetnames:
        if index in name.split('_'):
            ishave = True
            ws = w.get_sheet_by_name(name)
            col = ws.cell(row=3, column=1).value
            tablename = name.split('_')[1]
            ww = wb[tablename]
            colname = ww.cell(row=1, column=col).value
            print 'index name:%s' % index
            print u'该索引是基于表%s的%s列的' %(tablename, colname)
            break
    if not ishave:
        print u'该索引不存在'

# 测试
#wb = load_workbook('../data/test.xlsx')
#help_all()
# help_database('test')
# help_table('test', 'ccc')
# help_view('R1', 'test')
# HELP('help')