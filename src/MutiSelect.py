#!/usr/bin/env python
# -*- coding: utf-8 -*-

from prettytable import PrettyTable
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import xlrd
import re
from Select import *

#currentdb = 'test'

def JoinSelect(sql):
    sql = sql.lower()
    sqlItem = re.split(r'[, ]', sql)
    RemoveKong(sqlItem)
    # print sqlItem
    tables = sqlItem[sqlItem.index('from')+1 : sqlItem.index('where')]
    condition = sqlItem[sqlItem.index('where')+1 : ]
    targetcol = sqlItem[sqlItem.index('select')+1 : sqlItem.index('from')]
    key = condition[0].split('.')[1]
    #print key

    # 表格的对应列
    tabcol = []
    colname = []
    for i in tables:
        tabcol.append([])
    #print tabcol
    # print tables, condition, targetcol
    for i in targetcol:
        for j in xrange(len(tables)):
            if tables[j] == i.split('.')[0]:
                tabcol[j].append(i.split('.')[1])
                colname.append(i.split('.')[1])
    #print tabcol
    PrintTable = PrettyTable()
    PrintTable.field_names = colname
    #print PrintTable

    # 同时打开两个表，查找等值条件添加到PrintTable
    wb = load_workbook('data/'+currentdb+'.xlsx')
    try:
        ws1 = wb[tables[0]]
        ws2 = wb[tables[1]]
    except:
        print 'table not exist.'
        return 0
    #print ws1.title, ws2.title
    headers1 = []
    headers2 = []
    for cell in list(ws1.rows)[0]:
        headers1.append(cell.value)
    for cell in list(ws2.rows)[0]:
        headers2.append(cell.value)
    #print headers1,headers2
    keycolnum1 = headers1.index(key) + 1
    keycolnum2 = headers2.index(key) + 1

    colnum1 = []
    colnum2 = []
    for j in xrange(len(tabcol[0])):
        colnum1.append(headers1.index(tabcol[0][j]) + 1)
    for j in xrange(len(tabcol[1])):
        colnum2.append(headers2.index(tabcol[1][j]) + 1)
    #print colnum1
    #print colnum2


    #print keycolnum1,keycolnum2
    for i in xrange(2, ws1.max_row+1):
        keyval = ws1.cell(i,keycolnum1).value
        if keyval is None:
            i += 1
            continue
        for j in xrange(2, ws2.max_row+1):
            if ws2.cell(j, keycolnum2).value == keyval:
                tmp_row = []
                for k in colnum1:
                    tmp_row.append(ws1.cell(i, k).value)
                for k in colnum2:
                    tmp_row.append(ws2.cell(j,k).value)
                # print tmp_row
                PrintTable.add_row(tmp_row)

    print PrintTable

def UnionSelect(sql):
    sql = sql.lower()
    sql1 = sql.split('union select')[0]
    sql2 = sql.split('union select')[1]
    sql2 = 'select'+ sql2
    #print sql1,sql2
    res1 = Select(sql1, currentdb)
    res2 = Select(sql2, currentdb)
    # print res1
    # print res2
    # print PrettyTable
    sql = sql.lower()
    sqlItem = re.split(r'[, ]', sql)
    RemoveKong(sqlItem)
    colname = sqlItem[sqlItem.index('select')+1 : sqlItem.index('from')]
    # print colname
    # print sqlItem
    PrintTable = PrettyTable()
    PrintTable.field_names = colname
    for i in res1:
        PrintTable.add_row(i)
    for i in res2:
        if i not in res1:
            PrintTable.add_row(i)
    print 'Union Result:'
    print PrintTable

def QiantaoSelect(sql):
    sql = sql.lower()
    sqlItem = re.split(r'[, ]', sql)
    RemoveKong(sqlItem)
    print sqlItem

def MutiSelect(sql):
    if '.' in sql:
        JoinSelect(sql)
    elif 'union' in sql:
        UnionSelect(sql)
    elif '(' in sql and ')' in sql:
        QiantaoSelect(sql)


# 连接查询
#sql1 = 'SELECT aaa.name, aaa.course, aaa.grade stu.class stu.teacher FROM aaa, stu WHERE aaa.id = stu.id'
# 嵌套查询
#sql2 = 'SELECT Sname FROM Student WHERE Sno IN (SELECT Sno FROM SC WHERE Cno = 2);'
# 集合查询
#sql3 = 'select id, name from aaa where id = 14 union select id name from stu where age > 10'
#MutiSelect(sql1)
#MutiSelect(sql2)
#MutiSelect(sql3)