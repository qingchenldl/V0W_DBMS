#!/usr/bin/env python
# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import xlrd
import xlwt
from Select import CheckCondition

import os

def Deletedb(dbname):
    if os.path.exists('data/'+dbname+'.xlsx'):
        os.remove('data/'+dbname+'.xlsx')
        with open('data/DB.txt','r') as r:
            lines = r.readlines()
        with open('data/DB.txt','w') as w:
            for i in lines:
                if dbname not in i:
                    w.write(i)
        print "Delete Successfully."
    else:
        print "Database %s not exists" % dbname

def DeleteTable(currentdb, tablename):
    wb = load_workbook('data/' + currentdb + '.xlsx')
    sheets = wb.sheetnames
    # print sheets
    if tablename not in sheets:
        print 'The table %s not existed' % tablename
        return 0
    else:
        ws = wb[tablename]
        wb.remove(ws)
        ws = wb['Sheet']
        for i in xrange(1, ws.max_row):
            print ws.cell(row=i,column=1).value
            if ws.cell(row=i,column=1).value == tablename:
                for j in xrange(1, ws.max_column):
                    print ws.cell(row=i,column=j).value
                    ws.cell(row=i,column=j).value = ''
                break
        wb.save('data/' + currentdb + '.xlsx')
        print "The Table has been Delete."
# 删除表，写的不好，应该通过覆盖来删除。。

def DeleteCheckAndOr(currentdb, tablename, condition, key):
    wb = load_workbook('data/' + currentdb + '.xlsx')
    try:
        sheet = wb[tablename]
    except:
        print 'The table is not existed.'
        return 0

    headers = []
    for cell in list(sheet.rows)[0]:
        headers.append(cell.value)

    if key == 'and':
        andor_pos = condition.index('and')
    elif key == 'or':
        andor_pos = condition.index('or')

    AndCondition = []
    AndCondition.append(condition[0 : andor_pos])
    AndCondition.append(condition[andor_pos+1 : ])    # [['id', '<', '8'], ['age', '>', '10']]

    var1 = AndCondition[0][0]
    yunsuan1 = AndCondition[0][1]
    value1 = AndCondition[0][2]
    var2 = AndCondition[1][0]
    yunsuan2 = AndCondition[1][1]
    value2 = AndCondition[1][2]
    VarColnum1 = headers.index(var1) + 1
    VarColnum2 = headers.index(var2) + 1

    rows = []
    num = 0

    if key=='and':
        for i in xrange(2, sheet.max_row+1):
            tmp_row = []
            cell1 = sheet.cell(i, VarColnum1).value
            cell2 = sheet.cell(i, VarColnum2).value
            if CheckCondition(cell1, yunsuan1, value1) and CheckCondition(cell2, yunsuan2, value2):
                num += 1
            else:
                for j in xrange(1, sheet.max_column+1):
                    tmp_row.append(sheet.cell(i, j).value)
                    if tmp_row is not None:
                        rows.append(tmp_row)
    elif key == 'or':
        for i in xrange(2, sheet.max_row+1):
            tmp_row = []
            cell1 = sheet.cell(i, VarColnum1).value
            cell2 = sheet.cell(i, VarColnum2).value
            if CheckCondition(cell1, yunsuan1, value1) or CheckCondition(cell2, yunsuan2, value2):
                num += 1
            else:
                for j in xrange(1, sheet.max_column+1):
                    tmp_row.append(sheet.cell(i, j).value)
                    if tmp_row is not None:
                        rows.append(tmp_row)
    wb.remove(sheet)
    ns = wb.create_sheet(title=tablename)
    ns.append(headers)
    for i in xrange(0, len(rows)):
        if rows[i]:
            ns.append(rows[i])
    wb.save('data/' + currentdb + '.xlsx')
    print "%d rows have been deleted." % num

def DeleteCheckOne(currentdb, tablename, condition):
    wb = load_workbook('data/' + currentdb + '.xlsx')
    try:
        sheet = wb[tablename]
    except:
        print 'The table is not existed.'
        return 0

    headers = []
    for cell in list(sheet.rows)[0]:
        headers.append(cell.value)
    # print headers
    var = condition[0]
    yunsuan = condition[1]
    value = condition[2]
    VarColnum = headers.index(var) + 1

    rows = []
    num = 0         # 删除行的数目
    if yunsuan == '=':
        for i in xrange(2, sheet.max_row+1):
            tmp_row = []
            #print sheet.cell(i, VarColnum).value
            if sheet.cell(i, VarColnum).value == value:
                num += 1
            else:
                for j in xrange(1, sheet.max_column+1):
                    tmp_row.append(sheet.cell(i, j).value)
            if tmp_row is not None:
                rows.append(tmp_row)
            #print tmp_row

    elif yunsuan == '!=':
        for i in xrange(2, sheet.max_row+1):
            tmp_row = []
            #print sheet.cell(i, VarColnum).value
            if sheet.cell(i, VarColnum).value != value:
                num += 1
            else:
                for j in xrange(1, sheet.max_column+1):
                    tmp_row.append(sheet.cell(i, j).value)
            if tmp_row is not None:
                rows.append(tmp_row)

    elif yunsuan == '>':
        for i in xrange(2, sheet.max_row+1):
            tmp_row = []
            #print sheet.cell(i, VarColnum).value
            if sheet.cell(i, VarColnum).value > value and len(sheet.cell(i, VarColnum).value)>=len(value):
                num += 1
            else:
                for j in xrange(1, sheet.max_column+1):
                    tmp_row.append(sheet.cell(i, j).value)
            if tmp_row is not None:
                rows.append(tmp_row)

    elif yunsuan == '>=':
        for i in xrange(2, sheet.max_row+1):
            tmp_row = []
            #print sheet.cell(i, VarColnum).value
            if sheet.cell(i, VarColnum).value >= value and len(sheet.cell(i, VarColnum).value)>=len(value):
                num += 1
            else:

                for j in xrange(1, sheet.max_column+1):
                    tmp_row.append(sheet.cell(i, j).value)
            if tmp_row is not None:
                rows.append(tmp_row)

    elif yunsuan == '<':
        for i in xrange(2, sheet.max_row+1):
            tmp_row = []
            #print sheet.cell(i, VarColnum).value
            if sheet.cell(i, VarColnum).value < value and len(sheet.cell(i, VarColnum).value)<=len(value):
                num += 1
            else:

                for j in xrange(1, sheet.max_column+1):
                    tmp_row.append(sheet.cell(i, j).value)
            if tmp_row is not None:
                rows.append(tmp_row)

    elif yunsuan == '<=':
        for i in xrange(2, sheet.max_row+1):
            tmp_row = []
            #print sheet.cell(i, VarColnum).value
            if sheet.cell(i, VarColnum).value < value and len(sheet.cell(i, VarColnum).value)<=len(value):
                num += 1
            else:

                for j in xrange(1, sheet.max_column+1):
                    tmp_row.append(sheet.cell(i, j).value)
            if tmp_row is not None:
                rows.append(tmp_row)

    wb.remove(sheet)
    ns = wb.create_sheet(title=tablename)
    ns.append(headers)
    for i in xrange(0, len(rows)):
        if rows[i]:
            ns.append(rows[i])
    wb.save('data/' + currentdb + '.xlsx')
    print "%d rows have been deleted." % num

def DeleteColumn(currentdb, sql):
    sql = sql.lower()
    sqlItem = sql.split(' ')
    tablename = sqlItem[sqlItem.index('from')+1]

    if 'where' not in sqlItem:
        wb = load_workbook('data/' + currentdb + '.xlsx')
        sheet = wb[tablename]
        num = 0
        for i in xrange(2, sheet.max_row+1):
            num += 1
            for j in xrange(1, sheet.max_column+1):
                #print sheet.cell(row=i,column=j)
                sheet.cell(row=i,column=j).value = ''
        wb.save('data/'+currentdb+'.xlsx')
        print '%d lines have deleted.' % num
    else:
        condition = sqlItem[sqlItem.index('where')+1 : ]   # 利用select处理where
        # print condition
        if 'and' in condition:
            DeleteCheckAndOr(currentdb, tablename, condition, 'and')
        elif 'or' in condition:
            DeleteCheckAndOr(currentdb, tablename, condition, 'or')
        else:
            DeleteCheckOne(currentdb, tablename, condition)


#DeleteTable('test','bbb')
# sql = 'Delete * FROM aaa'
# DeleteColumn('test',sql)





