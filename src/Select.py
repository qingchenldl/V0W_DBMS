#!/usr/bin/env python
# -*- coding: utf-8 -*-

# 用于打印结果表
from prettytable import PrettyTable
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import xlrd
import re

def Checkin(wb, targetTab, target, condition, o):
    try:
        ws = wb[targetTab]
    except:
        print 'The table %s not exist' % targetTab
        return 1
    allvalue = Space.getAll(ws)
    if len(target) == 0:
        for i in xrange(ws.max_column):
            target.append(ws[chr(ord('A') + i) + '1'].value)
    try:
        result = []
        result.append(target)
        for i in xrange(len(allvalue)):
            flag = False
            for key in xrange(1, len(condition)):
                if condition[key] == allvalue[i][condition[0]]:
                    flag = True
            if flag:
                temp = []
                for tar in target:
                    temp.append(allvalue[i][tar])
                result.append(temp)
    except:
        return 0
    if len(o) != 0:
        i = 0
        for i in xrange(len(result[0])):
            if o[1] == result[0][i]:
                break
        k = result.pop(0)
        result = sorted(result, cmp=lambda x, y: cmp(x[i], y[i]))
        if o[0] == 'desc':
            result = result[::-1]
        result.insert(0, k)
    return result

def Checklike(wb, targetTab, target, condition, o):
    try:
        ws = wb[targetTab]
    except:
        print 'The table %s not exist' % targetTab
        return 1
    allvalue = Space.getAll(ws)
    if len(target) == 0:
        for i in xrange(ws.max_column):
            target.append(ws[chr(ord('A') + i) + '1'].value)
    try:
        result = []
        result.append(target)
        for i in xrange(len(allvalue)):
            match = re.match(condition[1].replace('%', '(.*)'), allvalue[i][condition[0]])
            if match:
                temp = []
                for tar in target:
                    temp.append(allvalue[i][tar])
                result.append(temp)
    except:
        return 0
    if len(o) != 0:
        i = 0
        for i in xrange(len(result[0])):
            if o[1] == result[0][i]:
                break
        k = result.pop(0)
        result = sorted(result, cmp=lambda x, y: cmp(x[i], y[i]))
        if o[0] == 'desc':
            result = result[::-1]
        result.insert(0, k)
    return result

def CheckBed(wb, targetTab, target, condition, o):
    try:
        ws = wb[targetTab]
    except:
        print 'The table %s not exist' % targetTab
        return 1
    allvalue = Space.getAll(ws)
    if len(target) == 0:
        for i in xrange(ws.max_column):
            target.append(ws[chr(ord('A') + i) + '1'].value)
    try:
        result = []
        result.append(target)
        begin = []
        for i in xrange(len(allvalue)):
            if allvalue[i][condition[0]] >= condition[1] and allvalue[i][condition[0]] < condition[2]:
                begin.append(i)
        for i in xrange(len(allvalue)):
            if i in begin:
                temp = []
                for tar in target:
                    temp.append(allvalue[i][tar])
                result.append(temp)
    except:
        return 0
    if len(o) != 0:
        i = 0
        for i in xrange(len(result[0])):
            if o[1] == result[0][i]:
                break
        k = result.pop(0)
        result = sorted(result, cmp=lambda x, y: cmp(x[i], y[i]))
        if o[0] == 'desc':
            result = result[::-1]
        result.insert(0, k)
    return result

def CheckCondition(cell, yunsuan, value):        # 仅判断TF
    if yunsuan == '=':
        if cell == value:
                return True
        return False
    elif yunsuan == '!=':
        if cell != value:
            return True
        return False
    elif yunsuan == '>':
        if cell > value and len(cell) >= len(value):
                return True
        return False
    elif yunsuan == '>=':
        if cell >= value and len(cell) >= len(value):
                return True
        return False
    elif yunsuan == '<=':
        if cell <= value and len(cell) <= len(value):
                return True
        return False
    elif yunsuan == '<':
        if cell < value and len(cell) <= len(value):
                return True
        return False

def CheckAndOr(currentdb, tablename, condition, target, PrintTable, key):
    wb = xlrd.open_workbook('data/'+currentdb+'.xlsx')
    sheet = wb.sheet_by_name(tablename)
    headers = sheet.row_values(0)

    if key == 'and':
        andor_pos = condition.index('and')
    elif key == 'or':
        andor_pos = condition.index('or')

    AndCondition = []
    AndCondition.append(condition[0:andor_pos])
    AndCondition.append(condition[andor_pos+1 : ])    # [['id', '<', '8'], ['age', '>', '10']]

    var1 = AndCondition[0][0]
    yunsuan1 = AndCondition[0][1]
    value1 = AndCondition[0][2]
    var2 = AndCondition[1][0]
    yunsuan2 = AndCondition[1][1]
    value2 = AndCondition[1][2]
    VarColnum1 = headers.index(var1)
    VarColnum2 = headers.index(var2)

    res = []

    if target == '*':
        target = headers
    #print target
    TargetColnum = []
    for t in target:
        TargetColnum.append(headers.index(t))

    if key=='and':
        for i in xrange(1, sheet.nrows):
            tmp_row = []
            cell1 = sheet.cell_value(i, VarColnum1)
            cell2 = sheet.cell_value(i, VarColnum2)
            if CheckCondition(cell1, yunsuan1, value1) and CheckCondition(cell2, yunsuan2, value2):
                for j in TargetColnum:
                    tmp_row.append(sheet.cell_value(i, j))
                res.append(tmp_row)
                PrintTable.add_row(tmp_row)
    elif key == 'or':
        for i in xrange(1, sheet.nrows):
            tmp_row = []
            cell1 = sheet.cell_value(i, VarColnum1)
            cell2 = sheet.cell_value(i, VarColnum2)
            if CheckCondition(cell1, yunsuan1, value1) or CheckCondition(cell2, yunsuan2, value2):
                for j in TargetColnum:
                    tmp_row.append(sheet.cell_value(i, j))
                res.append(tmp_row)
                PrintTable.add_row(tmp_row)

    if 'order' in condition and 'by' in condition:
        orderkey = condition[condition.index('by') + 1]
        PrintTable.sortby = orderkey
        if 'desc' in condition:
            PrintTable.reversesort = True
        print PrintTable
    else:
        print PrintTable
    return res

def CheckOne(currentdb, tablename, condition, target, PrintTable):
    wb = xlrd.open_workbook('data/' + currentdb + '.xlsx')
    sheet = wb.sheet_by_name(tablename)
    headers = sheet.row_values(0)
    var = condition[0]
    yunsuan = condition[1]
    value = condition[2]
    VarColnum = headers.index(var)

    res = []
    if target == '*':
        target = headers
    TargetColnum = []
    for t in target:
        TargetColnum.append(headers.index(t))

    if yunsuan == '=':
        for i in xrange(1, sheet.nrows):
            tmp_row = []
            if sheet.cell_value(i,VarColnum) == value:
                for j in TargetColnum:
                    tmp_row.append(sheet.cell_value(i, j))
                res.append(tmp_row)
                PrintTable.add_row(tmp_row)
        #print PrintTable

    if yunsuan == '>':
        for i in xrange(1, sheet.nrows):
            tmp_row = []
            if sheet.cell_value(i,VarColnum) > value and len(sheet.cell_value(i,VarColnum)) >= len(value):
                for j in TargetColnum:
                    tmp_row.append(sheet.cell_value(i, j))
                res.append(tmp_row)
                PrintTable.add_row(tmp_row)
        #print PrintTable

    if yunsuan == '>=':
        for i in xrange(1, sheet.nrows):
            tmp_row = []
            if sheet.cell_value(i,VarColnum) >= value and len(sheet.cell_value(i,VarColnum)) >= len(value):
                for j in TargetColnum:
                    tmp_row.append(sheet.cell_value(i, j))
                res.append(tmp_row)
                PrintTable.add_row(tmp_row)
        #print PrintTable

    if yunsuan == '<':
        for i in xrange(1, sheet.nrows):
            tmp_row = []
            if sheet.cell_value(i,VarColnum) < value and len(sheet.cell_value(i,VarColnum)) <= len(value):
                for j in TargetColnum:
                    tmp_row.append(sheet.cell_value(i, j))
                res.append(tmp_row)
                PrintTable.add_row(tmp_row)
        #print PrintTable

    if yunsuan == '<=':
        for i in xrange(1, sheet.nrows):
            tmp_row = []
            if sheet.cell_value(i,VarColnum) <= value and len(sheet.cell_value(i,VarColnum)) <= len(value):
                for j in TargetColnum:
                    tmp_row.append(sheet.cell_value(i, j))
                res.append(tmp_row)
                PrintTable.add_row(tmp_row)
        #print PrintTable
    if 'order' in condition and 'by' in condition:
        orderkey = condition[condition.index('by') + 1]
        PrintTable.sortby = orderkey
        if 'desc' in condition:
            PrintTable.reversesort = True
        print PrintTable
    else:
        print PrintTable
    return res

def RemoveSpace(alist):     # 去除列表元素中的空格
    for istr in alist:
        istr.replace(' ','')
    return alist

def RemoveKong(alist):           # 去除列表的空元素
    while '' in alist:
        alist.remove('')
    return alist

def Select(sql, currentdb):
    sql = sql.lower()
    sqlItem = re.split(r'[, ]',sql)
    RemoveKong(sqlItem)
    print sqlItem
    if 'where' in sqlItem:
        pos = []
        for i in range(0, len(sqlItem)):
            if sqlItem[i]=='select' or sqlItem[i]=='from' or sqlItem[i]=='where':
                pos.append(i)
        #print pos
        tablename = sqlItem[pos[1]+1]
        colname = sqlItem[pos[0]+1 : pos[1]]
        condition = sqlItem[pos[2]+1 : ]
        #print condition
        wb = xlrd.open_workbook('data/' + currentdb + '.xlsx')
        sheet = wb.sheet_by_name(tablename)  # 表
        if '*' in colname:
            colname = sheet.row_values(0)
        PrintTable = PrettyTable()
        PrintTable.field_names = colname

        # print tablename
        # print colname
        # print condition
        # print PrintTable

        if 'and' in condition:
            #condition.remove('and')
            res = CheckAndOr('test', tablename, condition,colname, PrintTable, 'and')

        elif 'or' in condition:
            res = CheckAndOr('test', tablename, condition, colname, PrintTable, 'or')

        else:
            res = CheckOne('test', tablename, condition,colname, PrintTable)

    elif 'where' not in sqlItem:
        pos = []
        for i in range(0, len(sqlItem)):
            if sqlItem[i] == 'select' or sqlItem[i] == 'from' or sqlItem[i] == 'where':
                pos.append(i)
        # print pos
        tablename = sqlItem[pos[1] + 1]
        colname = sqlItem[pos[0] + 1: pos[1]]
        RemoveKong(colname)

        # print colname

        wb = xlrd.open_workbook('data/' + currentdb + '.xlsx')
        sheet = wb.sheet_by_name(tablename)

        colnumlist = []
        if '*' in colname:
            colname = sheet.row_values(0)
            for c in xrange(len(colname)):
                colnumlist.append(c)
        else:
            headers = sheet.row_values(0)
            #print headers
            for c in colname:
                colnumlist.append(headers.index(c))
        #print colnumlist
        res = []
        PrintTable = PrettyTable()
        PrintTable.field_names = colname
        for i in xrange(1,sheet.nrows):
            tmp_row = []
            for j in colnumlist:
                tmp_row.append(sheet.cell_value(i,j))
            #print tmp_row
            res.append(tmp_row)

            PrintTable.add_row(tmp_row)

        if 'order' in sqlItem and 'by' in sqlItem:
            orderkey = sqlItem[sqlItem.index('by')+1]
            PrintTable.sortby = orderkey
            if 'desc' in sqlItem:
                PrintTable.reversesort = True
            print PrintTable
        else:
            print PrintTable
    return res

#sql = 'SELECT * FROM aaa WHERE id = 14 and age < 20 order by name DESC'
# sql = 'Select id course grade From aaa where id = 13'
#sql =
#Select(sql,'test')
