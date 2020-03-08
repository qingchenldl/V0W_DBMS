#!/usr/bin/env python
# -*- coding: utf-8 -*-

#currentdb = 'test'
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import BiTree

# 建库
def CreateDatabase(filename):
    file = open('data/DB.txt', 'a+')
    data = file.readlines()
    if filename + '\n' in data:
        print 'Database %s has existed' % filename
        return 0  # 判断数据库是否已经存在
    file.writelines(filename + '\n')
    wb = Workbook()
    wb.save('data/'+filename + '.xlsx')
    file.close()
    print 'Database %s create sucessfully.' % filename
    return wb

# 建表
def CreateTable(currentdb, tablename, col):
    wb = load_workbook('data/'+currentdb+'.xlsx')
    sheets = wb.sheetnames
    #print sheets
    if tablename in sheets:
        print 'The table %s has existed' % tablename
        return 0
    else:
        nt = wb.create_sheet()  # nt    new table
        wt = wb.worksheets[0]  # wt    work table
        nt.title = tablename
        for i in xrange(0, len(col)):
            print col[i].split(' ')
            print col[i].split(' ')[0]
            nt[chr(ord('A') + i) + '1'] = col[i].split(' ')[0]  # 第一行记录表的属性
        col.insert(0, tablename)
        wt.append(col)
        wb.save('data/'+ currentdb +'.xlsx')
        print "Table %s Create Sucessfully." % tablename

# 建立视图
def CreateView(name, currentdb, sqlsave):
    try:
        file = open('data/'+currentdb + 'View.txt', 'a+')
        for s in file:
            if name in s:
                print 'The View: %s has already exsited.' % name
                return 0
        file.writelines(name + ':' + sqlsave + '\n')
        print 'Create View: %s Successfully.' % name
        file.close()
    except:
        file = open('data/'+currentdb + 'View.txt', 'w')
        file.writelines(name + ':' + sqlsave + '\n')
        print 'Create View: %s Successfully.' % name
        file.close()

# 建立索引
def CreateIndex(currentdb, name, tablename, tablecol):
    wb = load_workbook('data/Index.xlsx')
    w = load_workbook('data/'+currentdb+'.xlsx')
    try:
        indextab = w['index']
        indextab.cell(1,indextab.max_colunm+1).value = name
    except:
        indextab = w.create_sheet()
        indextab.title = 'index'
        indextab.cell(1, indextab.max_colunm + 1).value = name
    table = w[tablename]
    ws = wb.create_sheet()
    ws.title = name
    allval = []
    row = []
    col = []
    index = Bitree.BTree()
    for i in xrange(1, table.max_column + 1):
        print table.cell(row=1, column=i).value
        if table.cell(row=1, column=i).value == tablecol:
            for j in xrange(2, table.max_row + 1):
                allval.append(table.cell(row=j, column=i).value)
                index.Insert(table.cell(row=j, column=i).value, j, i)
                row.append(j)
                col.append(i)
            break
    ws.append(allval)
    ws.append(row)
    ws.append(rol)
    wb.save('data/Index.xlsx')
    return index

# sql = 'Select id course grade From aaa'
# #ViewCreate(currentdb, 'IdGrade', sql)
# CreateView('IdGrade',currentdb,sql)