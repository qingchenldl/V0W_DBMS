#!/usr/bin/env python
# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import re
from Select import CheckCondition
from User import *

currentdb = 'test'

def UpdateCheckCondition(currentdb, tablename, condition, SetVaList):
    wb = load_workbook('data/' + currentdb + '.xlsx')
    try:
        sheet = wb[tablename]
    except:
        print 'The table is not existed.'
        return 0

    headers = []
    for cell in list(sheet.rows)[0]:
        headers.append(cell.value)
    var = condition[0]
    yunsuan = condition[1]
    value = condition[2]
    VarColnum = headers.index(var) + 1

    # set value
    SetVar = []
    Setvalue = []
    SetVarColnum = []
    for i in xrange(len(SetVaList)):
        if SetVaList[i] == '=':
            SetVar.append(SetVaList[i-1])
            Setvalue.append(SetVaList[i+1])

    print SetVar
    print Setvalue

    for sv in SetVar:
        SetVarColnum.append(headers.index(sv) + 1)
    print SetVarColnum

    # 参照完整性约束检查
    checklist = []
    cc = wb['Sheet']
    for i in xrange(1, cc.max_row + 1):
        if cc.cell(i, 1).value == tablename:
            for j in xrange(2, cc.max_column + 1):
                if cc.cell(i, j).value is not None:
                    checklist.append(cc.cell(i, j).value)
    print checklist

    num = 0

    if yunsuan == '=':
        for i in xrange(2, sheet.max_row+1):
            if sheet.cell(i,VarColnum).value == value:
                if Checking(sheet, checklist, Setvalue):
                    t = 0
                    num += 1
                    for j in SetVarColnum:
                        sheet.cell(i, j).value = Setvalue[t]
                        t += 1
    else:
        print 'error, check value equal.'
        return 1

    wb.save('data/' + currentdb + '.xlsx')
    print "%d rows have been changed." % num

def UpdateCheckConditionAnd(currentdb, tablename, condition, SetVaList, key):
    wb = load_workbook('data/' + currentdb + '.xlsx')
    try:
        sheet = wb[tablename]
    except:
        print 'The table is not existed.'
        return 0

    headers = []
    for cell in list(sheet.rows)[0]:
        headers.append(cell.value)

    if key == 'and':
        andor_pos = condition.index('and')
    elif key == 'or':
        andor_pos = condition.index('or')

    AndCondition = []
    AndCondition.append(condition[0 : andor_pos])
    AndCondition.append(condition[andor_pos+1 : ])    # [['id', '<', '8'], ['age', '>', '10']]

    var1 = AndCondition[0][0]
    yunsuan1 = AndCondition[0][1]
    value1 = AndCondition[0][2]
    var2 = AndCondition[1][0]
    yunsuan2 = AndCondition[1][1]
    value2 = AndCondition[1][2]
    VarColnum1 = headers.index(var1) + 1
    VarColnum2 = headers.index(var2) + 1
    # set value
    SetVar = []
    Setvalue = []
    SetVarColnum = []
    for i in xrange(len(SetVaList)):
        if SetVaList[i] == '=':
            SetVar.append(SetVaList[i - 1])
            Setvalue.append(SetVaList[i + 1])

    print SetVar
    print Setvalue

    for sv in SetVar:
        SetVarColnum.append(headers.index(sv) + 1)
    print SetVarColnum

    # 参照完整性约束检查
    checklist = []
    cc = wb['Sheet']
    for i in xrange(1, cc.max_row + 1):
        if cc.cell(i, 1).value == tablename:
            for j in xrange(2, cc.max_column + 1):
                if cc.cell(i, j).value is not None:
                    checklist.append(cc.cell(i, j).value)
    print checklist

    num = 0
    if key=='and':
        for i in xrange(2, sheet.max_row+1):
            cell1 = sheet.cell(i, VarColnum1).value
            cell2 = sheet.cell(i, VarColnum2).value
            if CheckCondition(cell1, yunsuan1, value1) and CheckCondition(cell2, yunsuan2, value2):
                num += 1
                if Checking(sheet, checklist, Setvalue):
                    t = 0
                    num += 1
                    for j in SetVarColnum:
                        sheet.cell(i, j).value = Setvalue[t]
                        t += 1

    elif key == 'or':
        for i in xrange(2, sheet.max_row+1):
            cell1 = sheet.cell(i, VarColnum1).value
            cell2 = sheet.cell(i, VarColnum2).value
            if CheckCondition(cell1, yunsuan1, value1) or CheckCondition(cell2, yunsuan2, value2):
                num += 1
                if Checking(sheet, checklist, Setvalue):
                    t = 0
                    num += 1
                    for j in SetVarColnum:
                        sheet.cell(i, j).value = Setvalue[t]
                        t += 1

    wb.save('data/' + currentdb + '.xlsx')
    print "%d rows have been changed." % num



def Update(sql):
    sql = sql.lower()
    sqlItem = re.split(r'[, ]', sql)
    print sqlItem
    tablename = sqlItem[sqlItem.index('update') + 1]
    SetVaList = sqlItem[sqlItem.index('set') + 1 : sqlItem.index('where')]
    condition = sqlItem[sqlItem.index('where') + 1 : ]

    print tablename
    print SetVaList
    print condition

    if 'and' in condition:
        UpdateCheckConditionAnd(currentdb, tablename, condition, SetVaList, 'and')
    else:
        UpdateCheckCondition(currentdb, tablename, condition, SetVaList)


sql = 'UPDATE aaa SET name = v0w id = 14 WHERE id = 13 and name = ldl'
# sql = 'UPDATE aaa SET name = v0w id = 13 WHERE id = 14'
Update(sql)