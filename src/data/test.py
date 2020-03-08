#!/usr/bin/env python
# -*- coding: utf-8 -*-
from openpyxl import *
import hashlib

def md5(ss):  # 密码使用md5加密，提高安全性
    m = hashlib.md5()
    m.update(ss)
    return m.hexdigest()

def Login(user, passwd):
    wb = load_workbook('user.xlsx')
    sheet = wb.active
    ret = []  # ret 用于保存是否是管理员、操作的数据库、所具有的权限
    for i in xrange(2, sheet.max_row):
        if sheet.cell(row=i, column=1).value == user:
            if sheet.cell(row=i, column=2).value == md5(passwd):
                for j in xrange(3, sheet.max_column + 1):
                    ret.append(sheet.cell(row=i, column=j).value)
                print ret  # 测试获取的权限
                print "Login Successfully.%s,Welcome!" % user
                return ret
            else:
                print "your password is wrong."
                return [0]
        else:
            print "user not exists" 
            return [0]

Login('admin','admin')