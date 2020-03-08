#!/usr/bin/env python
# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import re
from Select import *
from User import *

currentdb = 'test'
def Removefuhao(strlist):
    li = []
    for istr in strlist:
        MatchObj = re.match('\(\'.*', istr)
        if MatchObj:
            istr = istr.replace('(','')
            istr = istr.replace('\'', '')
        MatchObj = re.match('.*\'\)', istr)
        if MatchObj:
            istr = istr.replace('\'', '')
            istr = istr.replace(')', '')

        MatchObj = re.match('\(.*', istr)
        if MatchObj:
            istr = istr.replace('(', '')
        MatchObj = re.match('.*\)', istr)
        if MatchObj:
            istr = istr.replace(')', '')

        li.append(istr)
    return li

def InsertAll(currentdb, tablename, valuelist):
    wb = load_workbook('data/' + currentdb + '.xlsx')
    try:
        ws = wb[tablename]
    except:
        print 'The table is not existed.'
    # 参照完整性约束检查
    checklist = []
    cc = wb['Sheet']
    for i in xrange(1, cc.max_row + 1):
        if cc.cell(i, 1).value == tablename:
            for j in xrange(2, cc.max_column + 1):
                if cc.cell(i, j).value is not None:
                    checklist.append(cc.cell(i, j).value)
    print checklist
    values = Dividelist(valuelist, ws.max_column)
    print values
    tmprow = 0
    num = len(valuelist)/ws.max_column
    print num
    for i in xrange(ws.max_row+1, ws.max_row+1+num):
        if Checking(ws, checklist, values[tmprow]):
            for j in xrange(1, ws.max_column + 1):
                #print i,j,j-1+tmprow*ws.max_column
                ws.cell(i, j).value = valuelist[j-1+tmprow*ws.max_column]
        tmprow += 1

    wb.save('data/' + currentdb + '.xlsx')
    print 'Insert %d row.' % num

def InsertPart(currentdb, tablename, targetcol, valuelist):
    wb = load_workbook('data/' + currentdb + '.xlsx')
    try:
        ws = wb[tablename]
    except:
        print 'The table is not existed.'

    headers = []
    for cell in list(ws.rows)[0]:
        headers.append(cell.value)

    ChecklistNum = []
    ColNum = []
    for t in targetcol:
        if t in headers:
            ColNum.append(headers.index(t) + 1)
            ChecklistNum.append(headers.index(t))
        else:
            print "%s not in table." % t
            return 1
    print ColNum
    print ChecklistNum

    t = 0 # 标记列表值第几个（用于插入）
    num = len(valuelist) / len(targetcol)

    for i in xrange(ws.max_row+1, ws.max_row+1+num):

        for j in ColNum:
            ws.cell(i, j).value = valuelist[t]
            t += 1

    wb.save('data/' + currentdb + '.xlsx')
    print 'Insert %d row.' % num


def Insert(sql):
    sql = sql.lower()
    sqlItem = re.split(r'[, ]', sql)
    print sqlItem
    tablename = sqlItem[sqlItem.index('into') + 1]
    #print tablename
    valuessss = sqlItem[sqlItem.index('values') + 1 : ]
    ColtoChange = sqlItem[sqlItem.index('into')+2 : sqlItem.index('values')]

    RemoveKong(valuessss)
    values = Removefuhao(valuessss)
    print values
    RemoveKong(ColtoChange)
    ColtoChange = Removefuhao(ColtoChange)
    print ColtoChange

    if ColtoChange == []:       # 插入全部
        InsertAll(currentdb, tablename, values)
    else:
        InsertPart(currentdb, tablename,ColtoChange, values)




# sql = 'INSERT INTO aaa VALUES (14 sh 19 xxx 95 19 ssc 20 PE 96)'
# Insert(sql)